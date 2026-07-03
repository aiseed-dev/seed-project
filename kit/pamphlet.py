"""カタログ PDF(たねの森様式)から商品台帳 xlsx を生成する。

データ源は2系統を突き合わせる:
- 注文書ページ(品番・品名・単価の全品一覧) … 品名と価格の正
- 品種紹介ページ(7列×2段の段組) … 説明・科・内容量・蒔き時・収穫

段組は語の座標で列に分解する。品番(4桁)の位置を列の中心とし、
中心間の中点を列境界にする。装飾用の二重描画文字(オオククララ等)は除外。
"""

from __future__ import annotations

import argparse
import copy
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from statistics import median

import pdfplumber
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

import ledger

CODE_RE = re.compile(r"(?<!\d)(\d{4})(?!\d)")
PRICE_RE = re.compile(r"\s+([^¥]{1,60}?)\s*¥([\d,]+)")
CATEGORY_RE = re.compile(r"([^=＝]*?)\s*[=＝]バイオダイナミック")
LATIN_RE = re.compile(r"^[\x20-\x7E]+$")
FAMILY_RE = re.compile(r"(\S+科)\s*(.*)$")

MAX_CODE = 2999
BAND_GAP = 20  # 品番のy座標がこれ以上離れたら別の段
LINE_GAP = 4  # 語のy座標がこの範囲なら同じ行
DESC_GAP = 14  # 行間がこれ以上空いたら説明文の終わり
PHOTO_MARGIN = 40  # 次の段の写真領域を避ける余白


@dataclass
class Detail:
    """品種紹介ページの1セル。"""

    code: str
    en: str = ""
    name: str = ""
    family: str = ""
    amount: str = ""
    sow: str = ""
    harvest: str = ""
    desc: str = ""
    category: str = ""


@dataclass
class Entry:
    """注文書ページの1行。"""

    code: str
    name: str
    price: int
    detail: Detail | None = field(default=None)


def parse_order_line(line: str) -> list[tuple[str, str, int]]:
    """注文書の1行から (品番, 品名, 単価) を全て抜き出す。"""
    found = []
    for m in CODE_RE.finditer(line):
        code = m.group(1)
        if not (1 <= int(code) <= MAX_CODE):
            continue
        m2 = PRICE_RE.match(line, m.end())
        if not m2:
            continue
        name = m2.group(1).strip()
        if not name or re.search(r"\d{4}", name):
            continue
        found.append((code, name, int(m2.group(2).replace(",", ""))))
    return found


def _is_banner(text: str) -> bool:
    """二重描画の装飾文字(ババーーガガ…)か。4桁品番は対象外。"""
    return (
        len(text) >= 6
        and len(text) % 2 == 0
        and all(text[i] == text[i + 1] for i in range(0, len(text), 2))
    )


def _lines_of(words: list[dict]) -> list[tuple[float, str]]:
    """語を行(top, テキスト)に再構成する。"""
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    lines: list[tuple[float, list[str]]] = []
    for w in words:
        if lines and abs(w["top"] - lines[-1][0]) < LINE_GAP:
            lines[-1][1].append(w["text"])
        else:
            lines.append((w["top"], [w["text"]]))
    return [(top, " ".join(parts)) for top, parts in lines]


def classify_cell(code: str, lines: list[tuple[float, str]], category: str) -> Detail:
    """セル内の行を 英名・和名・科・蒔き時・収穫・説明 に振り分ける。"""
    d = Detail(code=code, category=category)
    desc: list[str] = []
    prev_top: float | None = None
    seen_family = False
    for top, text in lines:
        if text == code:
            continue
        if desc and prev_top is not None and top - prev_top > DESC_GAP:
            break  # 説明文の途中で行間が空いたら以降はページの別要素
        prev_top = top
        if m := FAMILY_RE.match(text):
            d.family, d.amount = m.group(1), m.group(2).strip()
            seen_family = True
        elif text.startswith(("蒔き時", "植えつけ")):
            d.sow = text
            seen_family = True
        elif text.startswith(("収 穫", "収穫", "花 期", "花期")):
            d.harvest = text
            seen_family = True
        elif not seen_family and LATIN_RE.match(text):
            d.en = f"{d.en} {text}".strip()
        elif not seen_family:
            d.name = f"{d.name} {text}".strip()
        else:
            desc.append(text)
    d.desc = "".join(desc)
    return d


def _page_category(text: str) -> str:
    for line in text.splitlines():
        if m := CATEGORY_RE.search(line):
            head = m.group(1)
            return re.sub(r"^[\x20-\x7E]+", "", head).strip()
    return ""


def parse_detail_page(page) -> list[Detail]:
    """品種紹介ページを列に分解して Detail を返す。"""
    category = _page_category(page.extract_text() or "")
    words = [w for w in page.extract_words() if not _is_banner(w["text"])]
    code_words = [
        w
        for w in words
        if re.fullmatch(r"\d{4}", w["text"]) and 1 <= int(w["text"]) <= MAX_CODE
    ]
    if len(code_words) < 4:
        return []

    code_words.sort(key=lambda w: w["top"])
    bands: list[list[dict]] = []
    for w in code_words:
        if bands and w["top"] - bands[-1][-1]["top"] < BAND_GAP:
            bands[-1].append(w)
        else:
            bands.append([w])

    details: list[Detail] = []
    for i, band in enumerate(bands):
        if len(band) < 2:
            continue
        band.sort(key=lambda w: w["x0"])
        centers = [(w["x0"] + w["x1"]) / 2 for w in band]
        gap = median(centers[j + 1] - centers[j] for j in range(len(centers) - 1))
        band_top = min(w["top"] for w in band)
        if i + 1 < len(bands):
            band_end = bands[i + 1][0]["top"] - PHOTO_MARGIN
        else:
            band_end = page.height
        for w, center in zip(band, centers, strict=True):
            cell_words = [
                cw
                for cw in words
                if abs((cw["x0"] + cw["x1"]) / 2 - center) < gap / 2 - 2
                and band_top - 2 <= cw["top"] < band_end
            ]
            detail = classify_cell(w["text"], _lines_of(cell_words), category)
            details.append(detail)
    return details


def parse(pdf_path: Path) -> list[Entry]:
    """PDF 全体から品目一覧(詳細つき)を得る。"""
    entries: dict[str, Entry] = {}
    details: dict[str, Detail] = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            order_hits = [
                hit for line in text.splitlines() for hit in parse_order_line(line)
            ]
            if len(order_hits) >= 20:  # 注文書ページ
                for code, name, price in order_hits:
                    entries.setdefault(code, Entry(code, name, price))
                continue
            for d in parse_detail_page(page):
                details.setdefault(d.code, d)
    for code, entry in entries.items():
        entry.detail = details.get(code)
    return sorted(entries.values(), key=lambda e: e.code)


EXTRA_HEADERS = ["英名", "科", "内容量", "蒔き時", "収穫"]

SHOP_DEFAULTS = {
    "店名": "たねの森",
    "住所": "",
    "電話": "042-982-5023",
    "メール": "info@tanenomori.org",
    "振込先": "",
    "税率": 0.10,
    "税込表示": "はい",
}


def _default_category(code: str) -> str:
    if code.startswith("2"):
        return "書籍"
    if code.startswith("1"):
        return "ハーブ・花"
    return "野菜"


def write_ledger(entries: list[Entry], out: Path) -> None:
    """抽出結果を商品台帳 xlsx に書き出す(標準9列+補助列)。"""
    ledger.new(out)
    wb: Workbook = load_workbook(out)
    ws = wb["商品"]
    base = len(ledger.ITEM_HEADERS)
    for offset, header in enumerate(EXTRA_HEADERS):
        cell = ws.cell(row=1, column=base + 1 + offset, value=header)
        cell.font = copy.copy(ws.cell(row=1, column=1).font)
        cell.fill = copy.copy(ws.cell(row=1, column=1).fill)
    for e in entries:
        d = e.detail or Detail(code=e.code)
        ws.append(
            [
                e.code,
                e.name,
                "",  # 種別は店が記入(登録品種でないことの確認も店側)
                e.price,
                d.desc,
                "",  # 生産地
                "",  # 発芽率
                "販売中",
                d.category or _default_category(e.code),
                d.en,
                d.family,
                d.amount,
                d.sow,
                d.harvest,
            ]
        )
    cfg = wb["設定"]
    labels = {cfg.cell(row=r, column=1).value: r for r in range(1, 10)}
    for label, value in SHOP_DEFAULTS.items():
        if label in labels:
            cfg.cell(row=labels[label], column=2, value=value)
    wb.save(out)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="カタログPDFから商品台帳を生成する")
    parser.add_argument("pdf", type=Path, help="カタログ PDF")
    parser.add_argument(
        "-o", "--out", type=Path, default=Path("台帳.xlsx"), help="出力する台帳 xlsx"
    )
    args = parser.parse_args(argv)
    entries = parse(args.pdf)
    if not entries:
        print("品目を抽出できませんでした", file=sys.stderr)
        return 1
    write_ledger(entries, args.out)
    no_detail = [e.code for e in entries if e.detail is None]
    with_detail = len(entries) - len(no_detail)
    print(f"{len(entries)} 品目を書き出しました: {args.out}")
    print(f"  詳細(説明・蒔き時など)あり: {with_detail} 件")
    if no_detail:
        print(f"  詳細なし: {len(no_detail)} 件 → {', '.join(no_detail)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
