"""商品台帳 xlsx から注文様式 xlsx を自動生成する。

品番は入力規則(ドロップダウン)で選択、品種名・単価は自動表示。
名前付きセルで読み取り位置を固定し、印刷すればそのまま FAX 注文書になる。
記入例シート付き。
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from ledger import Item, Shop, read_items, read_shop

LINES = 12  # 注文明細の行数
CUSTOMER_FIELDS = ["氏名", "住所", "電話", "メール"]
LIST_SHEET = "商品リスト"

_thin = Side(style="thin", color="A9A297")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HEAD_FILL = PatternFill("solid", start_color="E8E2D0")


def _order_sheet(
    ws: Worksheet,
    shop: Shop,
    n_items: int,
    example: bool = False,
    items: list[Item] | None = None,
) -> dict[str, str]:
    """注文書のレイアウトを描き、名前付きセルにする座標を返す。"""
    names: dict[str, str] = {}
    for col, width in zip("ABCDE", [12, 26, 10, 8, 12], strict=True):
        ws.column_dimensions[col].width = width

    ws["A1"] = f"ご注文書({shop.name} 宛)"
    ws["A1"].font = Font(bold=True, size=14)
    contact = " / ".join(x for x in [shop.tel, shop.email] if x)
    ws["A2"] = f"送付先: {contact}(FAX・郵送・メール添付のいずれでも結構です)"
    ws["A2"].font = Font(size=9)

    row = 4
    ws.cell(row=row, column=1, value="お客様情報").font = Font(bold=True)
    row += 1
    for field in CUSTOMER_FIELDS:
        label = ws.cell(row=row, column=1, value=field)
        label.fill = HEAD_FILL
        label.border = BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        entry = ws.cell(row=row, column=2)
        entry.border = BORDER
        names[f"注文_{field}"] = f"$B${row}"
        if example:
            samples = {
                "氏名": "種子 太郎",
                "住所": "〒000-0000 ○○県○○市○○ 4-5-6",
                "電話": "090-0000-0000",
                "メール": "taro@example.com",
            }
            entry.value = samples[field]
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="ご注文内容").font = Font(bold=True)
    row += 1
    header_row = row
    for col, header in enumerate(["品番", "品種名", "単価", "数量", "金額"], start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    names["注文明細開始"] = f"$A${header_row}"
    names["注文明細終了"] = f"$A${header_row + LINES}"

    for i in range(LINES):
        r = header_row + 1 + i
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BORDER
        ws.cell(
            row=r, column=2
        ).value = f'=IF($A{r}="","",VLOOKUP($A{r},{LIST_SHEET}!$A:$C,2,FALSE))'
        ws.cell(
            row=r, column=3
        ).value = f'=IF($A{r}="","",VLOOKUP($A{r},{LIST_SHEET}!$A:$C,3,FALSE))'
        ws.cell(row=r, column=5).value = f'=IF(OR($A{r}="",$D{r}=""),"",$C{r}*$D{r})'

    if example and items:
        ws.cell(row=header_row + 1, column=1, value=items[0].code)
        ws.cell(row=header_row + 1, column=4, value=2)
        if len(items) > 1:
            ws.cell(row=header_row + 2, column=1, value=items[1].code)
            ws.cell(row=header_row + 2, column=4, value=1)

    dv_code = DataValidation(
        type="list",
        formula1=f"{LIST_SHEET}!$A$2:$A${n_items + 1}",
        allow_blank=True,
        showErrorMessage=True,
    )
    dv_qty = DataValidation(
        type="whole", operator="greaterThan", formula1="0", allow_blank=True
    )
    dv_code.add(f"A{header_row + 1}:A{header_row + LINES}")
    dv_qty.add(f"D{header_row + 1}:D{header_row + LINES}")
    ws.add_data_validation(dv_code)
    ws.add_data_validation(dv_qty)

    note_row = header_row + LINES + 2
    tax_note = "税込" if shop.tax_included else "税別"
    ws.cell(
        row=note_row,
        column=1,
        value=f"※価格は{tax_note}です。送料・お支払方法は請求書にてご案内します。",
    ).font = Font(size=9)
    return names


def build(ledger_path: Path, out: Path) -> None:
    """台帳から注文様式 xlsx を生成する。"""
    shop = read_shop(ledger_path)
    items = [i for i in read_items(ledger_path) if i.stock == "販売中"]

    wb = Workbook()

    ws_list = wb.active
    assert ws_list is not None
    ws_list.title = LIST_SHEET
    ws_list.append(["品番", "品種名", "単価"])
    for item in items:
        ws_list.append([item.code, item.name, item.price])
    ws_list.sheet_state = "hidden"

    ws = wb.create_sheet("注文書")
    names = _order_sheet(ws, shop, len(items))
    for name, coord in names.items():
        wb.defined_names[name] = DefinedName(name, attr_text=f"注文書!{coord}")

    ws_ex = wb.create_sheet("記入例")
    _order_sheet(ws_ex, shop, len(items), example=True, items=items)

    wb.active = wb.index(ws)
    wb.save(out)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="台帳から注文様式xlsxを生成する")
    parser.add_argument("ledger", type=Path, help="商品台帳 xlsx")
    parser.add_argument(
        "-o", "--out", type=Path, default=Path("注文書.xlsx"), help="出力ファイル"
    )
    args = parser.parse_args(argv)
    build(args.ledger, args.out)
    print(f"作成しました: {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
