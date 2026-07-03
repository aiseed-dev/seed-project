"""商品台帳 xlsx の様式生成と読み取り。

台帳(店の手元)が正。カタログ・注文様式・請求書はすべてここから派生する。
様式は名前付きセルで読む(セル座標のハードコード禁止)。
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

KINDS = ["固定種", "在来種", "F1"]
STOCKS = ["販売中", "品切れ", "終了"]

ITEM_HEADERS = [
    "品番",
    "品種名",
    "種別",
    "価格",
    "説明",
    "生産地",
    "発芽率",
    "在庫状態",
    "カテゴリ",
]

SETTING_LABELS = [
    ("店名", "たねの店(サンプル)"),
    ("住所", "〒000-0000 ○○県○○市○○ 1-2-3"),
    ("電話", "000-000-0000"),
    ("メール", "info@example.com"),
    ("振込先", "○○銀行 ○○支店 普通 0000000 タネノミセ"),
    ("税率", 0.08),
    ("税込表示", "はい"),
]

SAMPLE_ITEMS = [
    (
        "A-001",
        "真黒茄子",
        "固定種",
        440,
        "濃黒紫色の中長なす。",
        "埼玉県",
        "80%以上",
        "販売中",
        "なす",
    ),
    (
        "A-002",
        "みず茄子",
        "在来種",
        470,
        "水分が多く生食もできる。",
        "大阪府",
        "75%以上",
        "販売中",
        "なす",
    ),
    (
        "B-001",
        "のらぼう菜",
        "在来種",
        380,
        "早春のとう立ち菜。",
        "東京都",
        "85%以上",
        "品切れ",
        "菜っ葉",
    ),
    (
        "B-002",
        "山東菜",
        "固定種",
        330,
        "漬け菜に向く半結球白菜。",
        "長野県",
        "80%以上",
        "販売中",
        "菜っ葉",
    ),
    (
        "C-001",
        "打木赤皮甘栗かぼちゃ",
        "固定種",
        500,
        "鮮やかな赤皮の早生種。",
        "石川県",
        "80%以上",
        "終了",
        "かぼちゃ",
    ),
]


@dataclass
class Item:
    """商品台帳の1行。"""

    code: str
    name: str
    kind: str
    price: int
    desc: str
    origin: str
    germination: str
    stock: str
    category: str


@dataclass
class Shop:
    """設定シートの店舗情報。"""

    name: str
    address: str
    tel: str
    email: str
    bank: str
    tax_rate: float
    tax_included: bool


class LedgerError(Exception):
    """台帳の様式が読み取れない。"""


def _define(wb: Workbook, name: str, sheet: str, coord: str) -> None:
    wb.defined_names[name] = DefinedName(name, attr_text=f"{sheet}!{coord}")


def _resolve(wb: Workbook, name: str) -> tuple[Worksheet, str]:
    if name not in wb.defined_names:
        raise LedgerError(f"名前付きセル「{name}」がありません")
    dest = list(wb.defined_names[name].destinations)
    if not dest:
        raise LedgerError(f"名前付きセル「{name}」の参照先がありません")
    sheet_name, coord = dest[0]
    return wb[sheet_name], coord.replace("$", "")


def new(path: Path, sample: bool = False) -> None:
    """商品台帳の雛形を作成する。"""
    wb = Workbook()

    ws = wb.active
    assert ws is not None
    ws.title = "商品"
    head_font = Font(bold=True)
    head_fill = PatternFill("solid", start_color="E8E2D0")
    for col, header in enumerate(ITEM_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = head_font
        cell.fill = head_fill
    widths = [10, 24, 10, 8, 40, 10, 12, 10, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + col - 1)].width = width
    _define(wb, "商品開始", "商品", "$A$1")

    dv_kind = DataValidation(
        type="list", formula1=f'"{",".join(KINDS)}"', allow_blank=True
    )
    dv_stock = DataValidation(
        type="list", formula1=f'"{",".join(STOCKS)}"', allow_blank=True
    )
    kind_col = ITEM_HEADERS.index("種別") + 1
    stock_col = ITEM_HEADERS.index("在庫状態") + 1
    dv_kind.add(f"{chr(ord('A') + kind_col - 1)}2:{chr(ord('A') + kind_col - 1)}500")
    dv_stock.add(f"{chr(ord('A') + stock_col - 1)}2:{chr(ord('A') + stock_col - 1)}500")
    ws.add_data_validation(dv_kind)
    ws.add_data_validation(dv_stock)

    if sample:
        for row_values in SAMPLE_ITEMS:
            ws.append(row_values)

    cfg = wb.create_sheet("設定")
    cfg.column_dimensions["A"].width = 12
    cfg.column_dimensions["B"].width = 48
    for row, (label, value) in enumerate(SETTING_LABELS, start=1):
        cfg.cell(row=row, column=1, value=label).font = head_font
        cfg.cell(row=row, column=2, value=value)
        _define(wb, label, "設定", f"$B${row}")

    wb.save(path)


def _row_to_item(headers: dict[str, int], values: tuple) -> Item:
    def get(name: str) -> object:
        idx = headers[name]
        return values[idx] if idx < len(values) else None

    def text(name: str) -> str:
        v = get(name)
        return str(v).strip() if v is not None else ""

    price_raw = get("価格")
    try:
        price = int(price_raw)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        raise LedgerError(f"価格が数値ではありません: {price_raw!r}") from None
    return Item(
        code=text("品番"),
        name=text("品種名"),
        kind=text("種別"),
        price=price,
        desc=text("説明"),
        origin=text("生産地"),
        germination=text("発芽率"),
        stock=text("在庫状態"),
        category=text("カテゴリ") or "その他",
    )


def read_items(path: Path) -> list[Item]:
    """台帳から商品一覧を読む(名前付きセル「商品開始」基準)。"""
    wb = load_workbook(path, data_only=True)
    ws, coord = _resolve(wb, "商品開始")
    start = ws[coord]
    header_row = start.row
    headers: dict[str, int] = {}
    for idx, cell in enumerate(ws[header_row]):
        if cell.value is not None:
            headers[str(cell.value).strip()] = idx
    missing = [h for h in ITEM_HEADERS if h not in headers]
    if missing:
        raise LedgerError(f"台帳に列がありません: {', '.join(missing)}")

    items: list[Item] = []
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code = values[headers["品番"]]
        if code is None or str(code).strip() == "":
            continue
        items.append(_row_to_item(headers, values))
    return items


def read_shop(path: Path) -> Shop:
    """台帳の設定シートから店舗情報を読む。"""
    wb = load_workbook(path, data_only=True)

    def value(name: str) -> object:
        ws, coord = _resolve(wb, name)
        return ws[coord].value

    def text(name: str) -> str:
        v = value(name)
        return str(v).strip() if v is not None else ""

    tax_raw = value("税率")
    try:
        tax_rate = float(tax_raw)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        raise LedgerError(f"税率が数値ではありません: {tax_raw!r}") from None
    return Shop(
        name=text("店名"),
        address=text("住所"),
        tel=text("電話"),
        email=text("メール"),
        bank=text("振込先"),
        tax_rate=tax_rate,
        tax_included=text("税込表示") == "はい",
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="商品台帳の雛形を作成する")
    parser.add_argument("path", type=Path, help="作成する台帳 xlsx のパス")
    parser.add_argument("--sample", action="store_true", help="サンプル商品を入れる")
    args = parser.parse_args(argv)
    if args.path.exists():
        print(f"既にあります: {args.path}", file=sys.stderr)
        return 1
    new(args.path, sample=args.sample)
    print(f"作成しました: {args.path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
