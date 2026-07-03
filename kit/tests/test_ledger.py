from pathlib import Path

import pytest
from openpyxl import load_workbook

import ledger


@pytest.fixture
def book(tmp_path: Path) -> Path:
    path = tmp_path / "台帳.xlsx"
    ledger.new(path, sample=True)
    return path


def test_new_has_named_cells(book: Path) -> None:
    wb = load_workbook(book)
    for name in [
        "商品開始",
        "店名",
        "住所",
        "電話",
        "メール",
        "振込先",
        "税率",
        "税込表示",
    ]:
        assert name in wb.defined_names


def test_new_has_dropdowns(book: Path) -> None:
    wb = load_workbook(book)
    formulas = [dv.formula1 for dv in wb["商品"].data_validations.dataValidation]
    assert any("固定種" in f for f in formulas)
    assert any("品切れ" in f for f in formulas)


def test_read_items(book: Path) -> None:
    items = ledger.read_items(book)
    assert len(items) == len(ledger.SAMPLE_ITEMS)
    first = items[0]
    assert first.code == "A-001"
    assert first.name == "真黒茄子"
    assert first.kind == "固定種"
    assert first.price == 440
    assert first.category == "なす"


def test_read_shop(book: Path) -> None:
    shop = ledger.read_shop(book)
    assert shop.name == "たねの店(サンプル)"
    assert shop.tax_rate == pytest.approx(0.08)
    assert shop.tax_included is True


def test_read_items_skips_blank_rows(book: Path) -> None:
    wb = load_workbook(book)
    ws = wb["商品"]
    ws.cell(row=20, column=2, value="品番なしの行")  # 品番が空 → 無視される
    wb.save(book)
    assert len(ledger.read_items(book)) == len(ledger.SAMPLE_ITEMS)


def test_read_items_requires_named_cell(tmp_path: Path) -> None:
    from openpyxl import Workbook

    path = tmp_path / "壊れ.xlsx"
    Workbook().save(path)
    with pytest.raises(ledger.LedgerError):
        ledger.read_items(path)
