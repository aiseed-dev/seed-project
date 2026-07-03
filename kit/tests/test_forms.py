from pathlib import Path

import pytest
from openpyxl import load_workbook

import forms
import ledger


@pytest.fixture
def order_form(tmp_path: Path) -> Path:
    book = tmp_path / "台帳.xlsx"
    ledger.new(book, sample=True)
    out = tmp_path / "注文書.xlsx"
    forms.build(book, out)
    return out


def test_named_cells(order_form: Path) -> None:
    wb = load_workbook(order_form)
    for name in ["注文_氏名", "注文_住所", "注文_電話", "注文_メール", "注文明細開始"]:
        assert name in wb.defined_names


def test_sheets(order_form: Path) -> None:
    wb = load_workbook(order_form)
    assert "注文書" in wb.sheetnames
    assert "記入例" in wb.sheetnames
    assert wb[forms.LIST_SHEET].sheet_state == "hidden"


def test_list_sheet_has_selling_items_only(order_form: Path) -> None:
    wb = load_workbook(order_form)
    ws = wb[forms.LIST_SHEET]
    codes = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "A-001" in codes
    assert "B-001" not in codes  # 品切れ
    assert "C-001" not in codes  # 終了


def test_code_dropdown(order_form: Path) -> None:
    wb = load_workbook(order_form)
    formulas = [dv.formula1 for dv in wb["注文書"].data_validations.dataValidation]
    assert any(forms.LIST_SHEET in f for f in formulas)


def test_example_sheet_filled(order_form: Path) -> None:
    wb = load_workbook(order_form)
    ws = wb["記入例"]
    values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert "種子 太郎" in values
    assert "A-001" in values
