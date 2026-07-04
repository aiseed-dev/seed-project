from pathlib import Path

from openpyxl import load_workbook

import inventory
import ledger


def test_inventory_from_stock_state(tmp_path: Path) -> None:
    book = tmp_path / "台帳.xlsx"
    ledger.new(book, sample=True)
    rows = {r["item_code"]: r["qty"] for r in inventory.inventory(book)}
    assert rows["A-001"] == 1  # 販売中
    assert rows["B-001"] == 0  # 品切れ
    assert rows["C-001"] == 0  # 終了


def test_inventory_prefers_count_column(tmp_path: Path) -> None:
    book = tmp_path / "台帳.xlsx"
    ledger.new(book, sample=True)
    wb = load_workbook(book)
    ws = wb["商品"]
    ws.cell(row=1, column=10, value="在庫数")
    ws.cell(row=2, column=10, value=42)  # A-001
    wb.save(book)
    rows = {r["item_code"]: r["qty"] for r in inventory.inventory(book)}
    assert rows["A-001"] == 42
