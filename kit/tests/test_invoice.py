import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

import forms
import invoice
import ledger

TODAY = datetime.date(2026, 7, 3)


@pytest.fixture
def env(tmp_path: Path) -> dict[str, Path]:
    book = tmp_path / "台帳.xlsx"
    ledger.new(book, sample=True)
    form = tmp_path / "注文書.xlsx"
    forms.build(book, form)
    inbox = tmp_path / "注文受信"
    inbox.mkdir()
    return {
        "ledger": book,
        "form": form,
        "inbox": inbox,
        "out": tmp_path / "請求書",
        "book": tmp_path / "注文台帳.xlsx",
    }


def _fill_order(
    form: Path,
    dest: Path,
    name: str = "種子 太郎",
    lines: list[tuple[str, object]] | None = None,
) -> None:
    """注文書に記入して dest に保存する(顧客の記入を再現)。"""
    wb = load_workbook(form)
    ws = wb["注文書"]

    def coord(defined: str) -> str:
        _, c = next(wb.defined_names[defined].destinations)
        return c.replace("$", "")

    ws[coord("注文_氏名")] = name
    ws[coord("注文_住所")] = "〒000-0000 ○○県○○市"
    ws[coord("注文_電話")] = "090-0000-0000"
    ws[coord("注文_メール")] = "taro@example.com"
    header_row = ws[coord("注文明細開始")].row
    if lines is None:
        lines = [("A-001", 2), ("A-002", 1)]
    for i, (code, qty) in enumerate(lines):
        ws.cell(row=header_row + 1 + i, column=1, value=code)
        ws.cell(row=header_row + 1 + i, column=4, value=qty)
    wb.save(dest)


def test_process_order(env: dict[str, Path]) -> None:
    _fill_order(env["form"], env["inbox"] / "注文1.xlsx")
    numbers, failed = invoice.process(
        env["ledger"], env["inbox"], env["out"], env["book"], today=TODAY
    )
    assert numbers == ["2026-0001"]
    assert failed == []
    assert (env["out"] / "請求書-2026-0001.xlsx").exists()
    assert (env["inbox"] / "処理済み" / "注文1.xlsx").exists()

    wb = load_workbook(env["book"])
    rows = list(wb["注文"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2  # 1明細1行
    assert rows[0][0] == "2026-0001"
    assert rows[0][6] == "A-001"
    assert rows[0][10] == 440 * 2  # 金額=台帳の単価×数量


def test_numbering_increments(env: dict[str, Path]) -> None:
    _fill_order(env["form"], env["inbox"] / "a.xlsx")
    _fill_order(env["form"], env["inbox"] / "b.xlsx", name="種子 花子")
    numbers, _ = invoice.process(
        env["ledger"], env["inbox"], env["out"], env["book"], today=TODAY
    )
    assert numbers == ["2026-0001", "2026-0002"]

    # 追加処理でも連番が続く
    _fill_order(env["form"], env["inbox"] / "c.xlsx")
    numbers, _ = invoice.process(
        env["ledger"], env["inbox"], env["out"], env["book"], today=TODAY
    )
    assert numbers == ["2026-0003"]


def test_invoice_totals_tax_included(env: dict[str, Path]) -> None:
    _fill_order(env["form"], env["inbox"] / "注文.xlsx", lines=[("A-001", 2)])
    invoice.process(env["ledger"], env["inbox"], env["out"], env["book"], today=TODAY)
    wb = load_workbook(env["out"] / "請求書-2026-0001.xlsx")
    values = [
        c.value for row in wb["請求書"].iter_rows() for c in row if c.value is not None
    ]
    assert "合計(税込)" in values
    assert 880 in values
    assert "種子 太郎 様" in values


def test_unreadable_goes_to_fail_dir(env: dict[str, Path]) -> None:
    bad = env["inbox"] / "壊れ.xlsx"
    bad.write_bytes(b"not an xlsx")
    _fill_order(env["form"], env["inbox"] / "正常.xlsx")
    numbers, failed = invoice.process(
        env["ledger"], env["inbox"], env["out"], env["book"], today=TODAY
    )
    assert numbers == ["2026-0001"]
    assert failed == ["壊れ.xlsx"]
    assert (env["inbox"] / "未処理" / "壊れ.xlsx").exists()


def test_unknown_code_goes_to_fail_dir(env: dict[str, Path]) -> None:
    _fill_order(env["form"], env["inbox"] / "不明品番.xlsx", lines=[("X-999", 1)])
    numbers, failed = invoice.process(
        env["ledger"], env["inbox"], env["out"], env["book"], today=TODAY
    )
    assert numbers == []
    assert failed == ["不明品番.xlsx"]
    assert not env["book"].exists()  # 記帳されない


def test_empty_lines_rejected(env: dict[str, Path]) -> None:
    _fill_order(env["form"], env["inbox"] / "明細なし.xlsx", lines=[])
    numbers, failed = invoice.process(
        env["ledger"], env["inbox"], env["out"], env["book"], today=TODAY
    )
    assert numbers == []
    assert failed == ["明細なし.xlsx"]
