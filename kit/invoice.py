"""受信した注文 xlsx を読み取り、注文台帳に記帳して請求書 xlsx を生成する。

- 注文書は名前付きセル(注文_氏名 など・注文明細開始)で読む
- 注文番号は「年-連番」を注文台帳から採番する
- 読み取れないファイルは「未処理」フォルダへ移す(黙って捨てない)
- 単価は台帳(正)から引く。注文書内の計算式は使わない
"""

from __future__ import annotations

import argparse
import datetime
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName

from ledger import Item, Shop, read_items, read_shop

BOOK_HEADERS = [
    "注文番号",
    "受付日",
    "氏名",
    "住所",
    "電話",
    "メール",
    "品番",
    "品種名",
    "数量",
    "単価",
    "金額",
]

_thin = Side(style="thin", color="A9A297")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HEAD_FILL = PatternFill("solid", start_color="E8E2D0")


class OrderError(Exception):
    """注文書が読み取れない。"""


@dataclass
class Line:
    item: Item
    qty: int

    @property
    def amount(self) -> int:
        return self.item.price * self.qty


@dataclass
class Order:
    customer: dict[str, str]  # 氏名・住所・電話・メール
    lines: list[Line]

    @property
    def total(self) -> int:
        return sum(line.amount for line in self.lines)


def _resolve(wb, name: str):
    if name not in wb.defined_names:
        raise OrderError(f"名前付きセル「{name}」がありません")
    dest = list(wb.defined_names[name].destinations)
    if not dest:
        raise OrderError(f"名前付きセル「{name}」の参照先がありません")
    sheet_name, coord = dest[0]
    return wb[sheet_name], coord.replace("$", "")


def read_order(path: Path, items: list[Item]) -> Order:
    """注文書 xlsx を読み取る。"""
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        raise OrderError(f"ファイルを開けません: {e}") from e

    customer: dict[str, str] = {}
    for field in ["氏名", "住所", "電話", "メール"]:
        ws, coord = _resolve(wb, f"注文_{field}")
        value = ws[coord].value
        customer[field] = str(value).strip() if value is not None else ""
    if not customer["氏名"]:
        raise OrderError("氏名が空です")

    by_code = {item.code: item for item in items}
    ws, coord = _resolve(wb, "注文明細開始")
    header_row = ws[coord].row
    _, end_coord = _resolve(wb, "注文明細終了")
    end_row = ws[end_coord].row
    headers: dict[str, int] = {}
    for idx, cell in enumerate(ws[header_row]):
        if cell.value is not None:
            headers[str(cell.value).strip()] = idx + 1
    for required in ("品番", "数量"):
        if required not in headers:
            raise OrderError(f"注文明細に「{required}」列がありません")

    lines: list[Line] = []
    for r in range(header_row + 1, end_row + 1):
        code_v = ws.cell(row=r, column=headers["品番"]).value
        qty_v = ws.cell(row=r, column=headers["数量"]).value
        code = str(code_v).strip() if code_v is not None else ""
        if not code:
            continue
        if code not in by_code:
            raise OrderError(f"台帳にない品番です: {code}")
        try:
            qty = int(qty_v)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            raise OrderError(f"数量が読めません(品番 {code}): {qty_v!r}") from None
        if qty <= 0:
            raise OrderError(f"数量が正の数ではありません(品番 {code}): {qty}")
        lines.append(Line(item=by_code[code], qty=qty))
    if not lines:
        raise OrderError("注文明細が空です")
    return Order(customer=customer, lines=lines)


def _open_book(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "注文"
    ws.append(BOOK_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
    wb.defined_names["注文開始"] = DefinedName("注文開始", attr_text="注文!$A$1")
    return wb


def next_number(book: Workbook, today: datetime.date) -> str:
    """注文番号(年-連番)を採番する。"""
    ws, coord = _resolve(book, "注文開始")
    header_row = ws[coord].row
    prefix = f"{today.year}-"
    max_serial = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_col=1, values_only=True):
        value = row[0]
        if value is None:
            continue
        text = str(value)
        if text.startswith(prefix):
            try:
                max_serial = max(max_serial, int(text[len(prefix) :]))
            except ValueError:
                continue
    return f"{today.year}-{max_serial + 1:04d}"


def record(book: Workbook, number: str, today: datetime.date, order: Order) -> None:
    """注文台帳に記帳する(1明細1行)。"""
    ws, _coord = _resolve(book, "注文開始")
    c = order.customer
    for line in order.lines:
        ws.append(
            [
                number,
                today.isoformat(),
                c["氏名"],
                c["住所"],
                c["電話"],
                c["メール"],
                line.item.code,
                line.item.name,
                line.qty,
                line.item.price,
                line.amount,
            ]
        )


def write_invoice(
    path: Path, shop: Shop, number: str, today: datetime.date, order: Order
) -> None:
    """請求書 xlsx を生成する。"""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "請求書"
    for col, width in zip("ABCDE", [12, 26, 10, 8, 12], strict=True):
        ws.column_dimensions[col].width = width

    ws["A1"] = "請求書"
    ws["A1"].font = Font(bold=True, size=16)
    ws["D1"] = f"注文番号: {number}"
    ws["D2"] = f"発行日: {today.isoformat()}"

    ws["A3"] = f"{order.customer['氏名']} 様"
    ws["A3"].font = Font(bold=True, size=12)
    ws["A4"] = order.customer["住所"]

    row = 6
    for col, header in enumerate(["品番", "品種名", "単価", "数量", "金額"], start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    for line in order.lines:
        row += 1
        values = [
            line.item.code,
            line.item.name,
            line.item.price,
            line.qty,
            line.amount,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER

    row += 1
    total = order.total
    rate = shop.tax_rate
    if shop.tax_included:
        inner_tax = total - round(total / (1 + rate))
        ws.cell(row=row, column=4, value="合計(税込)").font = Font(bold=True)
        cell = ws.cell(row=row, column=5, value=total)
        cell.font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=4, value=f"内消費税({rate:.0%})")
        ws.cell(row=row, column=5, value=inner_tax)
    else:
        tax = round(total * rate)
        ws.cell(row=row, column=4, value="小計")
        ws.cell(row=row, column=5, value=total)
        row += 1
        ws.cell(row=row, column=4, value=f"消費税({rate:.0%})")
        ws.cell(row=row, column=5, value=tax)
        row += 1
        ws.cell(row=row, column=4, value="合計").font = Font(bold=True)
        cell = ws.cell(row=row, column=5, value=total + tax)
        cell.font = Font(bold=True)

    row += 2
    ws.cell(row=row, column=1, value="お振込先").font = Font(bold=True)
    ws.cell(row=row + 1, column=1, value=shop.bank)
    ws.cell(
        row=row + 2,
        column=1,
        value="※送料は実費をご案内します。お支払いは商品到着後で結構です(後払い)。",
    ).font = Font(size=9)

    row += 4
    ws.cell(row=row, column=1, value=shop.name).font = Font(bold=True)
    contact = " / ".join(x for x in [shop.address, shop.tel, shop.email] if x)
    ws.cell(row=row + 1, column=1, value=contact).font = Font(size=9)

    wb.save(path)


def process(
    ledger_path: Path,
    inbox: Path,
    out: Path,
    book_path: Path,
    today: datetime.date | None = None,
) -> tuple[list[str], list[str]]:
    """inbox 内の注文 xlsx を一括処理する。(成功した注文番号, 未処理ファイル名)"""
    today = today or datetime.date.today()
    shop = read_shop(ledger_path)
    items = read_items(ledger_path)

    done_dir = inbox / "処理済み"
    fail_dir = inbox / "未処理"
    out.mkdir(parents=True, exist_ok=True)

    numbers: list[str] = []
    failed: list[str] = []
    for path in sorted(inbox.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        try:
            order = read_order(path, items)
        except OrderError as e:
            print(f"未処理: {path.name}: {e}", file=sys.stderr)
            fail_dir.mkdir(exist_ok=True)
            shutil.move(str(path), fail_dir / path.name)
            failed.append(path.name)
            continue
        book = _open_book(book_path)
        number = next_number(book, today)
        record(book, number, today, order)
        book.save(book_path)
        write_invoice(out / f"請求書-{number}.xlsx", shop, number, today, order)
        done_dir.mkdir(exist_ok=True)
        shutil.move(str(path), done_dir / path.name)
        numbers.append(number)
        print(f"記帳: {number} ({order.customer['氏名']} / {order.total:,}円)")
    return numbers, failed


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="注文xlsxを記帳し請求書を生成する")
    parser.add_argument("ledger", type=Path, help="商品台帳 xlsx")
    parser.add_argument(
        "--inbox", type=Path, default=Path("注文受信"), help="注文xlsxを置くフォルダ"
    )
    parser.add_argument(
        "--out", type=Path, default=Path("請求書"), help="請求書の出力先"
    )
    parser.add_argument(
        "--book", type=Path, default=Path("注文台帳.xlsx"), help="注文台帳 xlsx"
    )
    args = parser.parse_args(argv)
    if not args.inbox.is_dir():
        print(f"フォルダがありません: {args.inbox}", file=sys.stderr)
        return 1
    numbers, failed = process(args.ledger, args.inbox, args.out, args.book)
    print(f"処理 {len(numbers)} 件 / 未処理 {len(failed)} 件")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
