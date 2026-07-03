# SPDX-License-Identifier: AGPL-3.0-only
"""店舗スタッフ API: 取込・一括操作・申込み対応・エクスポート。"""

import io

import httpx
import psycopg
import pytest
from openpyxl import load_workbook

from tests.conftest import TEST_DSN, auth

CSV_TEXT = (
    "品種名,分類slug,種苗区分,種別,価格,数量表記,採種年,生産地,"
    "発芽率,種子消毒,説明,栽培メモ\n"
    "みやま小かぶ,root-veg,種,固定種,360,約200粒,2025,埼玉県,"
    "2026年6月現在 85%以上,無処理,肉質緻密な小かぶ。,秋まき推奨\n"
    "真黒茄子,fruit-veg,種,固定種,360,約30粒,2025,埼玉県,"
    "80%以上,無処理,濃黒紫の中長なす。,\n"
    "発芽率なし,leaf-veg,種,固定種,360,,,埼玉県,,,,\n"
)


def make_shop(owner: str = "staff1") -> str:
    with psycopg.connect(TEST_DSN, autocommit=True) as conn:
        conn.execute(
            "INSERT INTO shared.app_users (id, display_name) VALUES (%s, %s)"
            " ON CONFLICT (id) DO NOTHING",
            (owner, "たねの森スタッフ"),
        )
        row = conn.execute(
            "INSERT INTO shared.shops (slug, code, name, region, is_verified)"
            " VALUES ('tanenomori', 'TANE', 'たねの森', '埼玉県日高市', true)"
            " RETURNING id"
        ).fetchone()
        assert row is not None
        conn.execute(
            "INSERT INTO shared.shop_members (shop_id, user_id, role)"
            " VALUES (%s, %s, 'owner')",
            (row[0], owner),
        )
        return str(row[0])


def csv_file(text: str = CSV_TEXT) -> dict:
    return {"file": ("listings.csv", io.BytesIO(text.encode()), "text/csv")}


@pytest.mark.anyio
async def test_requires_shop_membership(client: httpx.AsyncClient) -> None:
    res = await client.get("/api/v1/shop/listings", headers=auth("outsider"))
    assert res.status_code == 403
    assert res.json()["code"] == "NOT_SHOP_STAFF"


@pytest.mark.anyio
async def test_csv_import_and_bulk(client: httpx.AsyncClient) -> None:
    make_shop()
    res = await client.post(
        "/api/v1/shop/listings/import", files=csv_file(), headers=auth("staff1")
    )
    assert res.status_code == 200, res.text
    result = res.json()
    # 全行が新品種 → proposed(マスタ提案を生成)。表示不足の行はエラー
    assert result["proposed"] == 2
    assert result["errors"] == 1
    assert "指定種苗の表示" in result["rows"][2]["detail"]

    # 店舗出品として指定種苗表示が補完されている
    res = await client.get("/api/v1/shop/listings", headers=auth("staff1"))
    listings = res.json()
    assert len(listings) == 2
    assert all(row["requires_seed_label"] for row in listings)
    assert listings[0]["label_seller_name"] == "たねの森"

    # 一括で停止 → closed
    ids = [row["id"] for row in listings]
    res = await client.post(
        "/api/v1/shop/listings/bulk",
        json={"ids": ids, "action": "close"},
        headers=auth("staff1"),
    )
    assert res.json() == {"updated": 2}
    res = await client.get("/api/v1/shop/listings", headers=auth("staff1"))
    assert {row["status"] for row in res.json()} == {"closed"}

    # 一括で価格変更
    res = await client.post(
        "/api/v1/shop/listings/bulk",
        json={"ids": ids, "action": "price", "price_yen": 400},
        headers=auth("staff1"),
    )
    res = await client.get("/api/v1/shop/listings", headers=auth("staff1"))
    assert {row["price_yen"] for row in res.json()} == {400}


@pytest.mark.anyio
async def test_shop_request_flow_and_deals_export(
    client: httpx.AsyncClient,
) -> None:
    shop_id = make_shop()
    # 担当名を登録(承諾担当者の記録に使う)
    res = await client.patch(
        "/api/v1/shop/me",
        json={"contact_label": "種苗部 田中"},
        headers=auth("staff1"),
    )
    assert res.json()["contact_label"] == "種苗部 田中"

    # 取込 → 公開中の出品を買い手がカートへ → 申込み
    await client.post(
        "/api/v1/shop/listings/import", files=csv_file(), headers=auth("staff1")
    )
    res = await client.get("/api/v1/shop/listings", headers=auth("staff1"))
    listing_id = res.json()[0]["id"]
    await client.put(
        f"/api/v1/cart/items/{listing_id}",
        json={"quantity": 3},
        headers=auth("buyer"),
    )
    res = await client.post(
        "/api/v1/requests",
        json={"provider_kind": "shop", "provider_id": shop_id},
        headers=auth("buyer"),
    )
    assert res.status_code == 201, res.text
    request_id = res.json()["id"]
    request_no = res.json()["request_no"]

    # 店舗宛一覧に載る(requested が上)
    res = await client.get("/api/v1/shop/requests", headers=auth("staff1"))
    assert res.json()[0]["request"]["id"] == request_id

    # スタッフが承諾 → 完了
    await client.patch(
        f"/api/v1/requests/{request_id}",
        json={"status": "accepted"},
        headers=auth("staff1"),
    )
    await client.patch(
        f"/api/v1/requests/{request_id}",
        json={"status": "completed"},
        headers=auth("buyer"),
    )

    # 成約台帳(xlsx): 申込番号・店舗コード・承諾担当者・金額
    res = await client.get(
        "/api/v1/shop/export",
        params={"kind": "deals", "format": "xlsx"},
        headers=auth("staff1"),
    )
    assert res.status_code == 200
    ws = load_workbook(io.BytesIO(res.content)).active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][:2] == ("申込番号", "店舗コード")
    assert rows[1][0] == request_no
    assert rows[1][1] == "TANE"
    assert rows[1][4] == "種苗部 田中"
    assert rows[1][9] == 360 * 3

    # 統計: 申込み数が載る
    res = await client.get("/api/v1/shop/stats", headers=auth("staff1"))
    stats = {row["listing_id"]: row["request_count"] for row in res.json()}
    assert stats[listing_id] == 1

    # CSV エクスポート(listings)
    res = await client.get(
        "/api/v1/shop/export",
        params={"kind": "listings", "format": "csv"},
        headers=auth("staff1"),
    )
    assert res.status_code == 200
    assert "みやま小かぶ" in res.content.decode("utf-8-sig")


@pytest.mark.anyio
async def test_patch_shop_owner_only(client: httpx.AsyncClient) -> None:
    shop_id = make_shop()
    with psycopg.connect(TEST_DSN, autocommit=True) as conn:
        conn.execute(
            "INSERT INTO shared.app_users (id, display_name) VALUES ('staff2', 'новый')"
            " ON CONFLICT (id) DO NOTHING"
        )
        conn.execute(
            "INSERT INTO shared.shop_members (shop_id, user_id, role)"
            " VALUES (%s, 'staff2', 'staff')",
            (shop_id,),
        )
    res = await client.patch(
        "/api/v1/shop",
        json={"return_policy": "未開封のみ7日以内"},
        headers=auth("staff2"),
    )
    assert res.status_code == 403
    res = await client.patch(
        "/api/v1/shop",
        json={"return_policy": "未開封のみ7日以内"},
        headers=auth("staff1"),
    )
    assert res.json()["return_policy"] == "未開封のみ7日以内"

    res = await client.get("/api/v1/shop/members", headers=auth("staff1"))
    assert len(res.json()) == 2
